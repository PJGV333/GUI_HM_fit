from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
import numpy as np


def write_results_xlsx(
    output_path: str | Path,
    *,
    constants: list[dict] | None = None,
    statistics: dict | None = None,
    results_text: str = "",
    export_data: dict | None = None,
) -> None:
    """
    Write an XLSX file with results.

    This is a local version of the results exporter.
    """
    from .processors import nmr_processor
    from openpyxl.styles import Font

    constants = constants or []
    statistics = statistics or {}
    export_data = export_data or {}

    output_path = Path(output_path)

    def as_dataframe(_name: str, data: Any, *, allow_none: bool = True) -> Optional[pd.DataFrame]:
        if data is None:
            return None if allow_none else pd.DataFrame()
        if isinstance(data, pd.DataFrame):
            return data
        if isinstance(data, pd.Series):
            return data.to_frame()
        if isinstance(data, dict):
            return pd.DataFrame(list(data.items()), columns=["key", "value"])
        return pd.DataFrame(data)

    def format_sheet(writer: pd.ExcelWriter, sheet_name: str, title: str):
        if sheet_name not in writer.sheets:
            return
        ws = writer.sheets[sheet_name]
        ws.insert_rows(1)
        ws.cell(row=1, column=1, value=title)
        
        # Merge cells across table width
        max_col = ws.max_column
        if max_col > 1:
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
        
        # Bold and slightly larger font for the title
        ws.cell(row=1, column=1).font = Font(bold=True, size=12)
        
        # Freeze panes so header (row 2) stays visible
        ws.freeze_panes = "A3"

    is_nmr_export = any(
        key in export_data for key in ("Chemical_Shifts", "Calculated_Chemical_Shifts", "signal_names")
    )

    TITLES_MAP = {
        "Constants": "Global constants and fitted parameters",
        "Statistics": "Summary of fit statistics",
        "Report": "Detail Report / Text Summary",
        "Model": "Model definition / Stoichiometry",
        "Selected_species": "Selected / observable species concentrations (M)",
        "All_species": "All species concentrations (M)",
        "Tot_con_comp": "Total component concentrations (M)",
        "Molar_Absortivities": "Molar absorptivities (a.u. or M^-1 cm^-1)",
        "Y_observed": "Observed spectra, Y_obs (a.u.)",
        "Y_raw": "Observed spectra before baseline correction",
        "Y_corrected": "Observed spectra after baseline correction",
        "Y_calculated": "Calculated spectra, Y_fit (a.u.)",
        "Baseline": "Per-spectrum baseline values",
        "Weights": "Per-wavelength weights used in residuals",
        "Stats": "Fit statistics / diagnostics",
        "Chemical_Shifts": "Observed chemical shifts (ppm)",
        "Calculated_Chemical_Shifts": "Calculated chemical shifts (ppm)",
        "Coefficients": "Calculated chemical shifts of species (ppm)",
        "K_calculated": "Calculated equilibrium constants",
        "Init_guess_K": "Initial guesses for equilibrium constants",
    }

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        # --- Common sheets ---
        if constants:
            df_constants = pd.DataFrame(constants)
            df_constants.to_excel(writer, sheet_name="Constants", index=False)
            format_sheet(writer, "Constants", TITLES_MAP["Constants"])

        if statistics:
            df_stats_common = pd.DataFrame(list(statistics.items()), columns=["metric", "value"])
            df_stats_common.to_excel(writer, sheet_name="Statistics", index=False)
            format_sheet(writer, "Statistics", TITLES_MAP["Statistics"])

        if results_text.strip():
            lines = [ln.rstrip("\n") for ln in str(results_text).splitlines()]
            pd.DataFrame({"text": lines}).to_excel(writer, sheet_name="Report", index=False)
            format_sheet(writer, "Report", TITLES_MAP["Report"])

        # --- Export payload sheets ---
        if is_nmr_export:
            modelo = as_dataframe("Model", export_data.get("modelo"), allow_none=False)
            C_T = as_dataframe("Tot_con_comp", export_data.get("C_T"), allow_none=False)
            dq = as_dataframe("Chemical_Shifts", export_data.get("Chemical_Shifts"), allow_none=False)
            dq_cal = as_dataframe(
                "Calculated_Chemical_Shifts",
                export_data.get("Calculated_Chemical_Shifts"),
                allow_none=False,
            )
            
            signal_names = list(export_data.get("signal_names") or [])
            column_names = list(export_data.get("column_names") or [])
            nas = list(export_data.get("non_absorbent_species") or [])

            # Correct data mapping: 
            # export_data["Co"] = All species, export_data["C"] = Selected species
            df_all = as_dataframe("All_species", export_data.get("Co"), allow_none=False)
            df_sel = as_dataframe("Selected_species", export_data.get("C"), allow_none=False)
            
            if not df_all.empty:
                df_all.index.name = "point"
                df_all.columns = [f"sp_{i+1}" for i in range(df_all.shape[1])]
            
            if not df_sel.empty:
                df_sel.index.name = "point"
                # Infer original indices for selected species
                n_all = df_all.shape[1] if not df_all.empty else 0
                if n_all > 0:
                    selected_indices = [i for i in range(n_all) if i not in nas]
                    if len(selected_indices) == df_sel.shape[1]:
                        df_sel.columns = [f"sp_{i+1}" for i in selected_indices]
                    else:
                        df_sel.columns = [f"sp_{i+1}" for i in range(df_sel.shape[1])]
                else:
                    df_sel.columns = [f"sp_{i+1}" for i in range(df_sel.shape[1])]

            if not C_T.empty:
                C_T.index.name = "point"
                if column_names and len(column_names) == C_T.shape[1]:
                    C_T.columns = column_names
                elif all(isinstance(c, (int, np.integer)) for c in C_T.columns):
                    C_T.columns = [f"ct_{i+1}" for i in range(C_T.shape[1])]

            if not dq.empty:
                dq.index.name = "point"
                if signal_names and len(signal_names) == dq.shape[1]:
                    dq.columns = signal_names
                elif all(isinstance(c, (int, np.integer)) for c in dq.columns):
                    dq.columns = [f"shift_obs_{i+1}" for i in range(dq.shape[1])]

            if not dq_cal.empty:
                dq_cal.index.name = "point"
                if signal_names and len(signal_names) == dq_cal.shape[1]:
                    dq_cal.columns = signal_names
                elif all(isinstance(c, (int, np.integer)) for c in dq_cal.columns):
                    dq_cal.columns = [f"shift_fit_{i+1}" for i in range(dq_cal.shape[1])]

            k_vals = export_data.get("k") or []
            percK = export_data.get("percK") or []
            k_names = [f"K{i+1}" for i in range(len(k_vals))]
            df_k = pd.DataFrame(
                {"Constants": list(k_vals), "Error (%)": list(percK)[: len(k_vals)]},
                index=k_names,
            )

            k_ini_vals = export_data.get("k_ini") or []
            k_ini_names = [f"K{i+1}" for i in range(len(k_ini_vals))]
            df_k_ini = pd.DataFrame({"Constants": list(k_ini_vals)}, index=k_ini_names)

            stats_table = export_data.get("stats_table")
            if stats_table:
                df_stats = pd.DataFrame(stats_table, columns=["metric", "Stats"]).set_index("metric")
            else:
                df_stats = pd.DataFrame(list(statistics.items()), columns=["metric", "Stats"]).set_index("metric")
            if "covfit" in df_stats.index:
                df_stats.loc["covfit", "Stats"] = str(df_stats.loc["covfit", "Stats"])

            coef = as_dataframe("Coefficients", export_data.get("Coefficients"), allow_none=True)
            if coef is None or coef.empty:
                try:
                    dq_arr = np.asarray(export_data.get("Chemical_Shifts"), dtype=float)
                    C_arr = np.asarray(export_data.get("C"), dtype=float)
                    C_T_arr = np.asarray(export_data.get("C_T"), dtype=float)

                    C_T_df = (
                        pd.DataFrame(C_T_arr, columns=column_names) if column_names else pd.DataFrame(C_T_arr)
                    )
                    D_cols, _ = nmr_processor.build_D_cols(
                        C_T_df, column_names, signal_names, default_idx=0
                    )

                    coef_mat = np.full((C_arr.shape[1], dq_arr.shape[1]), np.nan, dtype=float)
                    Xbase = np.asarray(C_arr, float)
                    finite_rows = np.isfinite(Xbase).all(axis=1)
                    nonzero_rows = np.linalg.norm(Xbase, axis=1) > 0.0
                    goodX = finite_rows & nonzero_rows
                    mask = np.isfinite(dq_arr) & np.isfinite(D_cols) & (np.abs(D_cols) > 0)

                    for j in range(dq_arr.shape[1]):
                        D = D_cols[:, j]
                        mj = mask[:, j] & goodX & np.isfinite(D) & (np.abs(D) > 0)
                        if int(mj.sum()) < 2:
                            continue
                        Xj = Xbase[mj, :] / D[mj][:, None]
                        y = dq_arr[mj, j]
                        coef_vec, *_ = np.linalg.lstsq(Xj, y, rcond=1e-10)
                        coef_mat[:, j] = coef_vec

                    coef = pd.DataFrame(coef_mat)
                except Exception:
                    coef = pd.DataFrame()

            if not isinstance(coef, pd.DataFrame):
                coef = pd.DataFrame(np.asarray(coef))
            
            if not coef.empty:
                if signal_names and len(signal_names) == coef.shape[1]:
                    coef.columns = signal_names
                elif all(isinstance(c, (int, np.integer)) for c in coef.columns):
                    coef.columns = [f"coef_{i+1}" for i in range(coef.shape[1])]
                coef.index.name = "species"

            sheets: Dict[str, pd.DataFrame] = {
                "Model": modelo if modelo is not None else pd.DataFrame(),
                "Selected_species": df_sel,
                "All_species": df_all,
                "Tot_con_comp": C_T if C_T is not None else pd.DataFrame(),
                "Chemical_Shifts": dq if dq is not None else pd.DataFrame(),
                "Calculated_Chemical_Shifts": dq_cal if dq_cal is not None else pd.DataFrame(),
                "Coefficients": coef if coef is not None else pd.DataFrame(),
                "K_calculated": df_k,
                "Init_guess_K": df_k_ini,
                "Stats": df_stats,
            }
            for name, df in sheets.items():
                df.to_excel(writer, sheet_name=name, index=True)
                format_sheet(writer, name, TITLES_MAP.get(name, name))
        else:
            if export_data:
                modelo = as_dataframe("Model", export_data.get("modelo"))
                if modelo is not None:
                    modelo.to_excel(writer, sheet_name="Model", index=False)
                    format_sheet(writer, "Model", TITLES_MAP["Model"])

                C = as_dataframe("Selected_species", export_data.get("C"))
                Co = as_dataframe("All_species", export_data.get("Co"))
                nas = list(export_data.get("non_abs_species") or [])

                if C is not None:
                    C.index.name = "point"
                    n_all = Co.shape[1] if Co is not None else 0
                    if n_all > 0:
                        selected_indices = [i for i in range(n_all) if i not in nas]
                        if len(selected_indices) == C.shape[1]:
                            C.columns = [f"sp_{i+1}" for i in selected_indices]
                        else:
                            C.columns = [f"sp_{i+1}" for i in range(C.shape[1])]
                    else:
                        C.columns = [f"sp_{i+1}" for i in range(C.shape[1])]
                    
                    C.to_excel(writer, sheet_name="Selected_species", index=True)
                    format_sheet(writer, "Selected_species", TITLES_MAP.get("Selected_species", "Selected species concentrations (M)"))

                # Wait, in spectroscopy, Co usually means concentrations of ALL species?
                # Let's check naming.
                Co = as_dataframe("All_species", export_data.get("Co"))
                if Co is not None:
                    Co.index.name = "point"
                    if all(isinstance(c, (int, np.integer)) for c in Co.columns):
                        Co.columns = [f"sp_{i+1}" for i in range(Co.shape[1])]
                    Co.to_excel(writer, sheet_name="All_species", index=True)
                    format_sheet(writer, "All_species", TITLES_MAP["All_species"])

                C_T = as_dataframe("Tot_con_comp", export_data.get("C_T"))
                if C_T is not None:
                    C_T.index.name = "point"
                    if all(isinstance(c, (int, np.integer)) for c in C_T.columns):
                        C_T.columns = [f"ct_{i+1}" for i in range(C_T.shape[1])]
                    C_T.to_excel(writer, sheet_name="Tot_con_comp", index=True)
                    format_sheet(writer, "Tot_con_comp", TITLES_MAP["Tot_con_comp"])

                A = export_data.get("A")
                nm = export_data.get("A_index") or export_data.get("nm")
                if A is not None:
                    dfA = as_dataframe("Molar_Absortivities", A, allow_none=True)
                    if dfA is not None:
                        if nm is not None:
                            dfA.index = nm
                        dfA.index.name = "nm"
                        if all(isinstance(c, (int, np.integer)) for c in dfA.columns):
                            dfA.columns = [f"A_{i+1}" for i in range(dfA.shape[1])]
                        dfA.to_excel(writer, sheet_name="Molar_Absortivities", index=True)
                        format_sheet(writer, "Molar_Absortivities", TITLES_MAP["Molar_Absortivities"])

                Y = export_data.get("Y")
                if Y is not None:
                    dfY = as_dataframe("Y_observed", Y, allow_none=True)
                    if dfY is not None:
                        if nm is not None:
                            dfY.index = nm
                        dfY.index.name = "nm"
                        if all(isinstance(c, (int, np.integer)) for c in dfY.columns):
                            dfY.columns = [f"yobs_{i+1}" for i in range(dfY.shape[1])]
                        dfY.to_excel(writer, sheet_name="Y_observed", index=True)
                        format_sheet(writer, "Y_observed", TITLES_MAP["Y_observed"])

                Y_raw = export_data.get("Y_raw")
                if Y_raw is not None:
                    dfY_raw = as_dataframe("Y_raw", Y_raw, allow_none=True)
                    if dfY_raw is not None:
                        if nm is not None:
                            dfY_raw.index = nm
                        dfY_raw.index.name = "nm"
                        if all(isinstance(c, (int, np.integer)) for c in dfY_raw.columns):
                            dfY_raw.columns = [f"yraw_{i+1}" for i in range(dfY_raw.shape[1])]
                        dfY_raw.to_excel(writer, sheet_name="Y_raw", index=True)
                        format_sheet(writer, "Y_raw", TITLES_MAP["Y_raw"])

                Y_corr = export_data.get("Y_corr")
                if Y_corr is not None:
                    dfY_corr = as_dataframe("Y_corrected", Y_corr, allow_none=True)
                    if dfY_corr is not None:
                        if nm is not None:
                            dfY_corr.index = nm
                        dfY_corr.index.name = "nm"
                        if all(isinstance(c, (int, np.integer)) for c in dfY_corr.columns):
                            dfY_corr.columns = [f"ycorr_{i+1}" for i in range(dfY_corr.shape[1])]
                        dfY_corr.to_excel(writer, sheet_name="Y_corrected", index=True)
                        format_sheet(writer, "Y_corrected", TITLES_MAP["Y_corrected"])

                yfit = export_data.get("yfit")
                if yfit is not None:
                    dfPhi = as_dataframe("Y_calculated", yfit, allow_none=True)
                    if dfPhi is not None:
                        if nm is not None:
                            dfPhi.index = nm
                        dfPhi.index.name = "nm"
                        if all(isinstance(c, (int, np.integer)) for c in dfPhi.columns):
                            dfPhi.columns = [f"ycal_{i+1}" for i in range(dfPhi.shape[1])]
                        dfPhi.to_excel(writer, sheet_name="Y_calculated", index=True)
                        format_sheet(writer, "Y_calculated", TITLES_MAP["Y_calculated"])

                baseline_vals = export_data.get("baseline_vals")
                if baseline_vals is not None:
                    df_base = pd.DataFrame({"baseline": np.asarray(baseline_vals, dtype=float).ravel()})
                    df_base.index.name = "point"
                    df_base.to_excel(writer, sheet_name="Baseline", index=True)
                    format_sheet(writer, "Baseline", TITLES_MAP["Baseline"])

                weights = export_data.get("weights")
                if weights is not None:
                    df_w = pd.DataFrame({"weight": np.asarray(weights, dtype=float).ravel()})
                    if nm is not None and len(nm) == len(df_w):
                        df_w.index = nm
                    df_w.index.name = "nm"
                    df_w.to_excel(writer, sheet_name="Weights", index=True)
                    format_sheet(writer, "Weights", TITLES_MAP["Weights"])

                k_vals = export_data.get("k") or []
                percK = export_data.get("percK") or []
                if k_vals:
                    names = [f"K{i+1}" for i in range(len(k_vals))]
                    df_k_spec = pd.DataFrame(
                        {"log10K": k_vals, "percK(%)": percK[: len(k_vals)]},
                        index=names,
                    )
                    df_k_spec.to_excel(writer, sheet_name="K_calculated")
                    format_sheet(writer, "K_calculated", TITLES_MAP["K_calculated"])

                k_ini = export_data.get("k_ini") or []
                if k_ini:
                    names_ini = [f"k{i+1}" for i in range(len(k_ini))]
                    df_k_ini_spec = pd.DataFrame({"init_guess": k_ini}, index=names_ini)
                    df_k_ini_spec.to_excel(writer, sheet_name="Init_guess_K")
                    format_sheet(writer, "Init_guess_K", TITLES_MAP["Init_guess_K"])
